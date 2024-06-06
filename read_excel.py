import openpyxl, os

#看當前目錄底下有哪下文件
os.listdir()

# file_path = os.getcwd() + "\\" + "test.xlsx"

#開啟excel
wbook = openpyxl.load_workbook('test.xlsx')

#查看所有sheet
# print(wbook.sheetnames())

"""
通過sheet_name取得工作表。
也可以用sheet_by_index(1), 通過index取得工作表。
"""
sheet1 = wbook['test_sheet1']
sheet2 = wbook['test_sheet2']

#row數量
#row整列的值
rnum = sheet1.iter_rows

#col數量
#col整欄的值
cnum = sheet1.iter_cols

#excel格定位取值
cellA2 = sheet1.cell(row=1, column=2).value
cellB3 = sheet1.cell(row=2, column=1).value

print(cellA2, cellB3)
# #col row 取值
# cellA3 = sheet1.row(2)[0].value
# cellC2 = sheet1.col(2)[1].value