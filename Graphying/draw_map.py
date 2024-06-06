import numpy as np
import openpyxl
import matplotlib.pyplot as plt

myWorkbook = openpyxl.load_workbook('map_data.xlsx')
map_data = myWorkbook['sheet2']
data_rows = map_data.max_row
data_columns = map_data.max_column
left_speed = np.array(
    [int(map_data.cell(row=i, column=1).value) for i in range(3, data_rows)])
# for row in range(2, data_rows):
#     left_speed = map_data.cell(row=row, column=1).value
right_speed = np.array(
    [int(map_data.cell(row=i, column=2).value) for i in range(3, data_rows)])
# for row in range(2, data_rows):
#     right_speed = map_data.cell(row=row, column=2).value

car_width = 10.3
global_x = 0
global_y = 0
car_direction = -np.pi / 2
communication_frequency = 0.1
n = 79  # lenth variable
j = 0  # point distance variable
real_x = []
real_y = []

# set graph style
plt.ion()
plt.figure(figsize=(6, 6))
plt.xlim([-10, 10])
plt.ylim([-10, 10])

for i in range(n):
    j += 0.1  # point distance
    car_velocity = (left_speed[i] + right_speed[i]) / 2
    car_angular_vel = (right_speed[i] - left_speed[i]) / car_width
    car_direction += -car_angular_vel * communication_frequency
    # print(left_speed, right_speed)
    '''
    cos  sin x
    -sin cos y
    '''
    car_x = car_velocity * np.cos(car_angular_vel) * 0.1
    car_y = car_velocity * np.sin(car_angular_vel) * 0.1
    global_x += np.cos(car_direction) * car_x + np.sin(car_direction) * car_y
    global_y += -np.sin(car_direction) * car_x + np.cos(car_direction) * car_y
    real_x.append(global_x)
    real_y.append(global_y)

    # plot point as time goes on
    plt.scatter(real_x, real_y, s=5, c='r', label="Fitting Curve", marker='o')
    # if em_filter :
    #     plt.plot(em_filter_x, em_filter_y, color = "C2", linewidth = 1)
    # if oe_filter :
    #     plt.plot(oe_filter_x, oe_filter_y, color = "C9", linewidth = 1)
    plt.ioff()
    plt.pause(0.005)
plt.show()
